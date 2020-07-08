'''
Code used to convert NoMorph.json to a spreadsheet
(Currently hard-coded to uses a spreadsheet with correct
 column names and NoMorph)
'''

import json
import openpyxl as pyxl

buildings = {}
with open('../bldgenergyuse/4 - Simulate buildings for urban morphologies/morphology_data/NoMorphEdited.json', 'r') as jsonFile:
    buildings = json.load(jsonFile)

# Code used to produce NoMorphEdited.json
#
# for id in buildings:
#     outputs = buildings[id]["outputs"]
#     newOutputs = {}
#     for key in outputs:
#         output = map(lambda x: {x[""]: {key:val for key, val in x.items() if key != "" }}, outputs[key])
#         buildings[id]["outputs"][key] = list(output)

filename="../processed_data/nomorph.xlsx"
wb = pyxl.load_workbook(filename)
wb.active = 0
sheet = wb.active


data = list(map(lambda x: x.value,list(sheet["A"])))
data = list(map(lambda x: x.replace(u'\xa0', ' '), data[:143]))

i = 0
row = 1
column = 2
for id in buildings:
    sheet.cell(row, column).value = int(id)
    row += 1
    for datum in data[1:4]:
        sheet.cell(row, column).value = buildings[id][datum]
        row += 1
    outputs = buildings[id]['outputs']
    row += 1
    curr_list = []
    curr_dict = {}
    for datum in data[5:]:
        if datum == 'area' or datum == 'end_uses' or datum == 'utilities_per_cond_area':
            curr_list = outputs[datum]
        else:
            if list(filter(lambda x: [*x][0] == datum, curr_list)) != []:
                curr_dict = list(filter(lambda x: [*x][0] == datum, curr_list))[0]
            else:
                sheet.cell(row, column).value = curr_dict[[*curr_dict][0]][datum]
        row += 1

    row = 1
    column += 1

wb.save(filename="../processed_data/nomorph.xlsx")
